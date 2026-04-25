"""
merge_results.py - 合并所有扫描器结果到单一 Excel（全量追加、去重、扫描时分秒）
Plan X: 支持只输出新岗位（--new-only）

修复记录 (2026-04-21):
  P1:  重复链接   → normalize_link() 标准化 URL 再去重
  P3:  score=0   → 过滤 score < 60 的岗位
  P4:  score字符串 → _int() 解析 P0/P1 字符串
  P8:  公司冗余   → extract_company() 从 URL 提取公司名
  P10: 排名无意义 → 去重+过滤后重新按 score 排序分配排名

用法：
  python merge_results.py              # 全量合并
  python merge_results.py --new-only   # 只输出今天新增的岗位
"""
import os, json, sys, re, argparse
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

WORKSPACE = Path(__file__).parent
RAW_DIR = WORKSPACE / "candidates" / "raw"
CONFIG_DIR = WORKSPACE / "config"
EXCEL_FILE = CONFIG_DIR / "HK_AI_Jobs_All.xlsx"

# Plan X
sys.path.insert(0, str(WORKSPACE / "scanners"))
from seen_jobs import load_seen_jobs, save_seen_jobs, check_job_status, update_job_entry

if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# ── helpers ────────────────────────────────────────────────────────────────

def load(path):
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return {"jobs": data}
        if isinstance(data, dict):
            return data
        return None
    except Exception:
        return None


# [P4] 修复: 支持 P0/P1 字符串解析
def _int(v, default=0):
    try:
        if isinstance(v, str):
            v = v.strip()
            if v.startswith("P") or v.startswith("p"):
                v = v[1:]   # 去掉 "P" 前缀
            v = v.strip()
        return int(v)
    except Exception:
        return default


def _str(v, default=""):
    if v is None:
        return default
    return str(v)


def get(j, *keys, default=None):
    for k in keys:
        v = j.get(k)
        if v is not None:
            return v
    return default if default is not None else ""


# [P1] 新增: 标准化链接，去除 query 参数差异
def normalize_link(link):
    """将链接标准化为 scheme://netloc/path，去除 query 和 fragment"""
    try:
        u = urlparse(link)
        return f"{u.scheme}://{u.netloc}{u.path}".rstrip("/")
    except Exception:
        # 回退：简单去除 query 参数
        return link.split("?")[0].rstrip("/")


# [P8] 新增: 从 URL 提取公司名
def extract_company(link, fallback=""):
    """从链接 URL 中提取公司名，fallback 为空则留空"""
    try:
        netloc = urlparse(link).netloc
        # Workday 平台: {company}.wd3.myworkdayjobs.com
        if "myworkdayjobs.com" in netloc:
            company = netloc.split(".")[0].upper()
            # 排除 generic 域名
            generic = {"www", "careers", "jobs", "apply", "en", "zh", "zh-tw", "zh-cn", "external", "search"}
            if company.lower() not in generic:
                return company
        # Taleo: {company}.taleo.net
        if ".taleo.net" in netloc:
            parts = netloc.split(".")[:-2]   # 去掉 taleo.net
            if parts:
                return parts[0].upper()
        # 一般域名: 取第一段
        first = netloc.split(".")[0].lower()
        generic = {"www", "careers", "jobs", "apply", "search", "career"}
        if first not in generic:
            return first.upper()
    except Exception:
        pass
    return fallback


# ── main ──────────────────────────────────────────────────────────────────

def merge(new_only=False):
    print(f"[MERGE] 扫描: {RAW_DIR}")
    if new_only:
        print("[MERGE] 模式: 只输出新岗位 (Plan X)")

    raw_files = sorted(RAW_DIR.glob("*_raw_????-??-??.json"))
    matched_files = sorted(RAW_DIR.glob("[a-z]*_????-??-??.json"))

    # [P1] 使用标准化链接去重: normalized_link -> (job, json_path)
    source_jobs = {}
    all_raw = list(raw_files) + list(matched_files)
    seen_sources = {}

    for fpath in all_raw:
        data = load(fpath)
        if not data:
            continue
        jobs = data.get("jobs", [])
        if not jobs:
            continue
        source = data.get("source",
                          re.sub(r"[_-]20\d\d-\d\d-\d\d.*", "", fpath.stem))
        seen_sources[source] = seen_sources.get(source, 0) + len(jobs)

        for j in jobs:
            link = get(j, "link", "url", "href", default="")
            if not link:
                continue
            json_path = str(fpath.relative_to(WORKSPACE))
            norm = normalize_link(link)

            if norm in source_jobs:
                existing, existing_path = source_jobs[norm]
                old_score = _int(get(existing, "score", "priority", default=0))
                new_score = _int(get(j, "score", "priority", default=0))
                if new_score > old_score:
                    source_jobs[norm] = (j, json_path)
            else:
                source_jobs[norm] = (j, json_path)

    jobs = [job for job, _ in source_jobs.values()]
    link_to_path = {norm: path for norm, (_, path) in source_jobs.items()}
    link_to_orig = {norm: list(source_jobs.keys())[list(
        normalize_link(get(j2, "link", "url", "href", default="")) for j2, _ in source_jobs.values()
    ).index(norm)] if False else norm for norm in source_jobs}  # 保持 norm 作为 key

    all_jobs = jobs
    print(f"\n[INFO] 发现 {len(all_raw)} 个 JSON 文件，{len(jobs)} 个去重岗位（链接标准化后）")
    for src, cnt in sorted(seen_sources.items()):
        print(f"       {src}: {cnt} 个")

    if not jobs:
        print("[WARN] 没有数据")
        return

    # ── Plan X: 加载 seen_jobs ─────────────────────────────────────────────────
    seen_data = load_seen_jobs()

    if new_only:
        new_jobs = []
        for j in jobs:
            link = get(j, "link", "url", "href", default="")
            title = get(j, "title", "Title", default="")
            status = check_job_status(link, title, seen_data)
            if status == "new":
                new_jobs.append(j)
        jobs = new_jobs
        print(f"[INFO] Plan X 过滤: {len(jobs)} 个新岗位（今天首次出现）")
        if not jobs:
            save_seen_jobs(seen_data)
            return

    # ── [P3] 过滤 score < 60 的岗位 ─────────────────────────────────────────────
    # 同时建立 normalized_link -> original_link 映射（用于查询 JSON 路径）
    norm_to_orig = {}
    for j in all_jobs:
        orig_link = get(j, "link", "url", "href", default="")
        if orig_link:
            norm_to_orig[normalize_link(orig_link)] = orig_link

    jobs = [j for j in jobs if _int(get(j, "score", "priority", default=0)) >= 60]
    print(f"[INFO] [P3] score>=60 过滤后: {len(jobs)} 个岗位")

    if not jobs:
        print("[WARN] 过滤后没有数据")
        return

    # ── Excel 读写 ────────────────────────────────────────────────────────

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(str(EXCEL_FILE))
        ws = wb.active
        existing_norm = set()
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row, 5).value   # col 5 = 链接
            if v:
                existing_norm.add(normalize_link(str(v)))
        print(f"[INFO] 现有 Excel: {ws.max_row - 1} 行")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        existing_norm = set()

    # 表头
    headers = [
        "编号", "来源平台", "公司", "职位", "链接", "发布时间",
        "匹配度", "匹配/不匹配原因", "JD总结", "JD路径", "扫描时间"
    ]
    widths = [5, 14, 18, 35, 42, 12, 8, 42, 52, 40, 16]

    if ws.max_row == 1:
        hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hf_f = Font(bold=True, color="FFFFFF", size=10)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf_f; c.fill = hf
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

    # 样式
    p0  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    p1  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    p2  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    nf  = Font(size=9)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    wrap = Alignment(wrap_text=True, vertical="top")

    def make_summary(jd_text):
        if not jd_text:
            return "JD信息不足"
        text = str(jd_text)
        lines = [l.strip() for l in text.split("\n")
                 if l.strip() and len(l.strip()) > 15]
        return ("\n".join(lines[:6]))[:800] if lines else text[:500]

    # [P10] 按扫描时间降序（新到旧），新职位排在前面
    jobs.sort(key=lambda x: _str(get(x, "scraped_at", "_scraped_at", "update_time"), ""), reverse=True)

    new_rows = 0
    next_rank = ws.max_row
    for j in jobs:
        orig_link = get(j, "link", "url", "href", default="")
        if not orig_link:
            continue
        norm_link = normalize_link(orig_link)
        if norm_link in existing_norm:
            continue

        score = _int(get(j, "score", "priority", "score_value", default=0))
        fill = p0 if score >= 75 else p1 if score >= 60 else p2

        post_date = _str(get(j, "post_date", "posted_at", "published_date", "date"))
        post_date = post_date[:10] if post_date else ""

        scraped_at = _str(get(j, "scraped_at", "_scraped_at", "update_time"))
        scraped_at = scraped_at[:16] if scraped_at else datetime.now().strftime("%Y-%m-%d %H:%M")

        jd_full = get(j, "full_jd", "jd_full", "description", default="")
        json_path = link_to_path.get(norm_link, "")

        match_reason = get(j, "match_reason", "_match_reason", "reason", "comment", default="")
        if not match_reason:
            reason_val = get(j, "scoringDetails", "score_details", default="")
            if reason_val:
                match_reason = _str(reason_val)[:200]

        priority = get(j, "priority", "level", default="")
        if priority and not match_reason:
            match_reason = f"[{priority}] {match_reason}"

        # [P8] 从 URL 提取公司名
        company = extract_company(orig_link, _str(get(j, "company", "Company"), ""))

        next_rank += 1
        row_data = [
            next_rank - 1,                                      # 1 排名
            _str(get(j, "_source", "source"), ""),             # 2 来源平台
            company,                                            # 3 公司（从 URL 提取）
            _str(get(j, "title", "Title"), ""),                # 4 职位
            orig_link,                                          # 5 链接
            post_date,                                           # 6 发布时间（留空）
            score,                                               # 7 匹配度
            _str(match_reason)[:300],                           # 8 匹配原因
            make_summary(jd_full),                               # 9 JD总结
            json_path,                                          # 10 JD路径
            scraped_at,                                         # 11 扫描时间
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=next_rank, column=col, value=val)
            c.font = nf; c.fill = fill; c.alignment = wrap; c.border = thin
        ws.row_dimensions[next_rank].height = 80
        existing_norm.add(norm_link)
        new_rows += 1

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(EXCEL_FILE))

    # ── Plan X: 更新 seen_jobs ─────────────────────────────────────────────────
    for j in all_jobs:
        link = get(j, "link", "url", "href", default="")
        title = get(j, "title", "Title", default="")
        company = get(j, "company", "Company", "source", default="")
        jd_text = get(j, "description", "jd", "full_jd", default="")
        status = check_job_status(link, title, seen_data)
        update_job_entry(link, title, company, jd_text, seen_data, status)
    save_seen_jobs(seen_data)
    print(f"[OK] seen_jobs.json 已更新 ({len(seen_data.get('jobs', {}))} 个已记录岗位)")

    total = ws.max_row - 1
    print(f"\n[OK] 合并完成")
    print(f"[OK] 新增 {new_rows} 行，总计 {total} 行（含表头）")
    print(f"[OK] 文件: {EXCEL_FILE}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="合并扫描结果到 Excel")
    parser.add_argument("--new-only", action="store_true",
                        help="只输出今天新增的岗位（Plan X）")
    args = parser.parse_args()
    merge(new_only=args.new_only)
