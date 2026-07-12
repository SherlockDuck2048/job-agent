"""
job_scanner_base.py - 通用扫描基础库
所有平台扫描脚本共用此基础库
"""
import json
import time
import random
import os
import sys
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# Fix Windows console encoding issue
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

CDP_URL = "http://127.0.0.1:9222"
WORKSPACE = r"C:\Users\ClawAdmin\.qclaw\workspace\job-agent"
RAW_DIR = os.path.join(WORKSPACE, "candidates", "raw")
DATE = datetime.now().strftime("%Y-%m-%d")

# AI 关键词过滤
AI_KEYWORDS = [
    'artificial intelligence', 'machine learning', 'deep learning',
    'genai', 'generative ai', 'llm', 'large language model',
    'data scientist', 'data science', 'nlp', 'natural language',
    'ai product', 'ai manager', 'ai analyst', 'ai engineer',
    'ai consultant', 'ai solution', 'ai strategy', 'ai director',
    'ai business', 'ai project', 'ai lead', 'ai specialist',
    'rag', 'retrieval', 'agent framework', 'prompt engineer',
    'computer vision', 'ml engineer', 'mlops',
]

EXCLUDE_KEYWORDS = [
    'air cargo', 'air freight', 'airline', 'flight',
    'intern', '實習', '实习', 'summer intern',
]

def is_ai_job(title: str, snippet: str = '') -> bool:
    """判断是否为 AI 相关岗位"""
    text = (title + ' ' + snippet).lower()
    # 排除非AI岗位
    for kw in EXCLUDE_KEYWORDS:
        if kw.lower() in text:
            return False
    # 匹配AI关键词
    for kw in AI_KEYWORDS:
        if kw in text:
            return True
    return False

def random_delay(min_s=1.5, max_s=3.5):
    """随机延迟,模拟人工操作"""
    time.sleep(random.uniform(min_s, max_s))

def clean_jd_text(text: str) -> str:
    """
    清洗 JD 文本，去除导航栏、按钮等无关内容
    """
    if not text:
        return ''
    
    # 常见无关内容模式
    noise_patterns = [
        r'Skip to main content',
        r'Manulife and John Hancock Careers',
        r'English\s*Sign In\s*Search for Jobs',
        r'page is loaded',
        r'Apply\s*locations\s*Hong Kong\s*time type',
        r'Full time\s*posted on\s*Posted',
        r'Days Ago\s*time left to apply',
        r'job requisition id',
        r'Return to Search Results',
        r'Similar Jobs',
        r'Connect with Us',
        r'Copyright \d{4}.*Manulife',
        r'Privacy Policy|Terms of Use|Accessibility',
    ]
    
    import re
    cleaned = text
    for pattern in noise_patterns:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
    
    # 清理多余空白
    cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
    cleaned = re.sub(r'[ \t]+', ' ', cleaned)
    
    return cleaned.strip()


def _strip_html(text: str) -> str:
    """去除 HTML 标签，保留纯文本"""
    import re
    # 移除 HTML 注释
    text = re.sub(r'<!--.*?-->', '', text, flags=re.DOTALL)
    # 移除 <style> 和 <script> 块
    text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.IGNORECASE | re.DOTALL)
    # 替换块级标签为空格，防止单词粘连
    text = re.sub(r'<(?:p|div|br|h[1-6]|li|tr)[^>]*>', ' ', text, flags=re.IGNORECASE)
    # 移除所有剩余 HTML 标签
    text = re.sub(r'<[^>]+>', '', text)
    # 解码 HTML 实体
    text = text.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"').replace('&#39;', "'").replace('\xa0', ' ')
    return text


def _split_paragraphs(text: str) -> list:
    """同时兼容 \n\n 和 \n 两种段落分隔符，返回非空段落列表"""
    import re
    # 先按 \n\n 分，再用 \n 进一步切分（合并行内换行）
    blocks = re.split(r'\n{2,}', text)
    paras = []
    for block in blocks:
        # 每块内按单换行拆分成句子，合并短句
        lines = [l.strip() for l in re.split(r'\n', block) if l.strip()]
        merged = ' '.join(lines)
        if merged:
            paras.append(merged)
    return paras


def _call_llm(system_prompt: str, user_content: str, max_tokens: int = 500) -> str:
    """
    调用 QClaw 本地 LLM API (OpenAI 兼容格式)
    Returns: LLM 响应文本，失败返回空字符串
    """
    import urllib.request as _urllib_request
    import json as _json

    base_url = os.environ.get('QCLAW_LLM_BASE_URL', 'http://127.0.0.1:19000/proxy/llm')
    api_key = os.environ.get('QCLAW_LLM_API_KEY', '')
    url = f'{base_url}/chat/completions'

    payload = {
        'model': 'default',
        'messages': [
            {'role': 'system', 'content': system_prompt},
            {'role': 'user', 'content': user_content},
        ],
        'max_tokens': max_tokens,
        'temperature': 0.3,
    }

    data = _json.dumps(payload, ensure_ascii=False).encode('utf-8')
    req = _urllib_request.Request(url, data=data, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('Authorization', f'Bearer {api_key}')

    try:
        resp = _urllib_request.urlopen(req, timeout=90)
        body = _json.loads(resp.read())
        content = body.get('choices', [{}])[0].get('message', {}).get('content', '')
        # Strip <think>...</think> tags from reasoning models
        if content:
            import re as _re
            content = _re.sub(r'<think[\s\S]*?</think\s*>', '', content).strip()
        return content
    except Exception:
        return ''


def extract_jd_summary(full_jd: str, max_chars: int = 500, use_llm: bool = True) -> str:
    """
    从完整 JD 中提取道法术器结构化摘要

    默认调用 LLM 生成「道→法→术→器→核心竞争力」五段摘要；
    LLM 失败时 fallback 到规则提取（关键词段落匹配）。

    Args:
        full_jd: JD 原始文本（可能含 HTML）
        max_chars: 摘要最大字符数（LLM 模式建议 ≥300，规则模式 300）
        use_llm: 是否使用 LLM 生成（默认 True）
    """
    if not full_jd:
        return ''

    # ── 预处理：去除 HTML + 清洗噪音 ────────────────────────────────────────
    plain = _strip_html(full_jd)
    cleaned = clean_jd_text(plain)

    # 截断过长 JD（避免 LLM token 浪费）
    jd_for_llm = cleaned[:4000] if len(cleaned) > 4000 else cleaned

    # ── LLM 模式：道法术器结构化摘要 ────────────────────────────────────────
    if use_llm and len(jd_for_llm) > 100:
        system_prompt = (
            "你是一个职位分析专家。根据职位描述，用「道法术器」框架生成结构化摘要。\n"
            "格式要求（严格遵循，每段一句话）：\n"
            "道：岗位的核心理念与价值追求。\n"
            "法：主要职责与工作方法。\n"
            "术：所需技能与专业能力。\n"
            "器：使用的工具、系统或资质。\n"
            "核心竞争力：该岗位最稀缺的复合能力。\n"
            "规则：只基于 JD 内容提取，不脑补；输出纯文本，无多余解释。"
        )
        llm_result = _call_llm(system_prompt, jd_for_llm, max_tokens=1200)
        if llm_result and ('道：' in llm_result or '道:' in llm_result):
            return llm_result.strip()[:max_chars]

    # ── 规则 fallback：关键词段落提取 ────────────────────────────────────────
    import re
    paragraphs = _split_paragraphs(cleaned)

    KEYWORD_PATTERNS = [
        r'responsibilities?[:\s]',
        r'duties?[:\s]',
        r'what you( will)? do[s]?[:\s]',
        r'job purpose[:\s]',
        r'position overview[:\s]',
        r'about the role[:\s]',
    ]
    for para in paragraphs:
        for pat in KEYWORD_PATTERNS:
            if re.search(pat, para, re.IGNORECASE):
                summary = para[:max_chars]
                if len(para) > max_chars:
                    summary = summary.rsplit(' ', 1)[0] + '...'
                return summary

    # 跳过前 2 段，取第 3 段
    if len(paragraphs) >= 3 and len(paragraphs[2]) > 80:
        summary = paragraphs[2][:max_chars]
        if len(paragraphs[2]) > max_chars:
            summary = summary.rsplit(' ', 1)[0] + '...'
        return summary

    # 取第一个有意义段落
    meaningful = [p for p in paragraphs if len(p) > 50]
    if meaningful:
        summary = meaningful[0][:max_chars]
        if len(meaningful[0]) > max_chars:
            summary = summary.rsplit(' ', 1)[0] + '...'
        return summary

    # 兜底
    summary = cleaned.strip()[:max_chars]
    if len(cleaned.strip()) > max_chars:
        summary = summary.rsplit(' ', 1)[0] + '...'
    return summary


def save_results(platform: str, jobs: list, status: str, notes: str = ''):
    """保存扫描结果到标准 JSON"""
    filename = f"{platform.lower().replace(' ', '_')}_{DATE}.json"
    filepath = os.path.join(RAW_DIR, filename)

    result = {
        "platform": platform,
        "scan_date": DATE,
        "scan_time": datetime.now().strftime("%H:%M:%S"),
        "total_found": len(jobs),
        "jobs": jobs,
        "status": status,
        "notes": notes
    }

    os.makedirs(RAW_DIR, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"✅ {platform}: {len(jobs)} 个岗位 → {filepath}")
    return filepath

def connect_browser():
    """连接 CDP 浏览器"""
    p = sync_playwright().start()
    try:
        browser = p.chromium.connect_over_cdp(CDP_URL)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        print(f"✅ 已连接 CDP: {CDP_URL}")
        return p, browser, context
    except Exception as e:
        p.stop()
        raise ConnectionError(f"❌ 无法连接 CDP ({CDP_URL}): {e}")

def new_page(context):
    """创建新标签页"""
    page = context.new_page()
    page.set_default_timeout(30000)
    return page

def safe_goto(page, url, wait='domcontentloaded', retries=2):
    """安全跳转,带重试"""
    for i in range(retries):
        try:
            page.goto(url, wait_until=wait, timeout=30000)
            return True
        except PlaywrightTimeout:
            print(f"  ⚠️ 加载超时 (尝试 {i+1}/{retries}): {url[:60]}")
            if i < retries - 1:
                random_delay(2, 4)
    return False

def scroll_to_bottom(page, max_scrolls=5):
    """滚动到底部加载更多内容"""
    for _ in range(max_scrolls):
        prev_height = page.evaluate("document.body.scrollHeight")
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        random_delay(1.5, 2.5)
        new_height = page.evaluate("document.body.scrollHeight")
        if new_height == prev_height:
            break

def extract_text(page, selector, default=''):
    """安全提取文本"""
    try:
        el = page.query_selector(selector)
        return el.inner_text().strip() if el else default
    except:
        return default

def extract_texts(page, selector):
    """安全提取多个文本"""
    try:
        els = page.query_selector_all(selector)
        return [el.inner_text().strip() for el in els if el.inner_text().strip()]
    except:
        return []

def extract_attr(page, selector, attr, default=''):
    """安全提取属性"""
    try:
        el = page.query_selector(selector)
        return el.get_attribute(attr) or default if el else default
    except:
        return default

def get_full_jd(page, url, jd_selectors=None):
    """进详情页获取完整 JD"""
    if not url or url.startswith('#'):
        return ''

    try:
        detail_page = new_page(page.context)
        if not safe_goto(detail_page, url):
            detail_page.close()
            return ''

        random_delay(1, 2)

        # 尝试点击展开按钮
        for expand_sel in ['button:has-text("See more")', 'button:has-text("Show more")',
                           '[data-tracking-control-name="public_jobs_show-more-html-btn"]']:
            try:
                btn = detail_page.query_selector(expand_sel)
                if btn:
                    btn.click()
                    random_delay(0.5, 1)
            except:
                pass

        # 尝试多个 JD 选择器
        selectors = jd_selectors or [
            '.job-description', '#job-description', '.jobsearch-jobDescriptionText',
            '[data-testid="job-description"]', '.description__text',
            '.job-details', '.jd-content', 'article', '.content'
        ]

        for sel in selectors:
            try:
                el = detail_page.query_selector(sel)
                if el:
                    text = el.inner_text().strip()
                    if len(text) > 100:
                        detail_page.close()
                        return text
            except:
                pass

        # 最后尝试 body
        text = detail_page.evaluate("document.body.innerText")
        detail_page.close()
        return text[:3000] if text else ''

    except Exception as e:
        print(f"  ⚠️ 详情页抓取失败: {str(e)[:50]}")
        try:
            detail_page.close()
        except:
            pass
        return ''


# ─────────────────────────────────────────────────────────────────────────────
# Plan C: 统一 JD 抓取函数(平台适配器模式)
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# Excel 追加函数：扫描器只追加自己的数据到 Excel
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_link(link: str) -> str:
    """
    归一化 URL，用于去重键。
    - 移除 query string（?ref=xxx）和 fragment（#xxx）
    - 移除末尾斜杠
    - 示例: https://linkedin.com/jobs/view/123?ref=ABC → https://linkedin.com/jobs/view/123
    """
    if not link:
        return link
    from urllib.parse import urlparse, urlunparse
    try:
        parsed = urlparse(link.strip())
        normalized = urlunparse((parsed.scheme, parsed.netloc.lower(),
                                  parsed.path.rstrip('/'), '', '', ''))
        return normalized
    except Exception:
        return link.strip().rstrip('/')


def _parse_scraped_ts(ts_str):
    """解析 scraped_at 字段，返回可比较的时间字符串（YYYY-MM-DD HH:MM），失败返回 ''。"""
    if not ts_str:
        return ''
    if isinstance(ts_str, (int, float)):
        return datetime.fromtimestamp(ts_str).strftime('%Y-%m-%d %H:%M')
    s = str(ts_str).strip()
    # 常见格式直接截取
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M',
                '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(s[:len(fmt.replace(' ', '')) if '%H' not in fmt
                                    else min(len(s), 16)], fmt).strftime('%Y-%m-%d %H:%M')
        except Exception:
            pass
    # ISO 格式带 Z
    try:
        return datetime.strptime(s[:19], '%Y-%m-%dT%H:%M:%S').strftime('%Y-%m-%d %H:%M')
    except Exception:
        pass
    return s[:16]


def append_scanner_to_excel(json_path: str, excel_path: str = None):
    """
    将单个扫描器的 JSON 结果追加到 Excel。

    双键去重：
      1. 归一化 URL（strip query/fragment）—— 同一职位不同 ref 参数视为相同
      2. Title + Company 组合键——跨平台同一岗位只保留一次

    同一岗位多次出现时，保留 scraped_at 最新版本（覆盖旧行）。
    """
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        print("[Excel] openpyxl not installed, skipping Excel update")
        return

    if excel_path is None:
        excel_path = os.path.join(WORKSPACE, "config", "HK_AI_Jobs_All.xlsx")

    if not os.path.exists(json_path):
        print(f"[Excel] JSON not found: {json_path}")
        return

    # ── 1. 加载扫描器 JSON ──────────────────────────────────────
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        jobs    = data.get('jobs', [])
        platform = data.get('source', 'Unknown')
    except Exception as e:
        print(f"[Excel] JSON read error: {e}")
        return

    if not jobs:
        print(f"[Excel] {platform}: no jobs, skip")
        return

    # ── 2. 读取 Excel 已有数据，构建去重索引 ────────────────────
    headers = ["ID", "Platform", "Company", "Title", "Link", "Location",
               "Priority", "Score", "Match Reason", "Post Date", "Scraped At",
               "JD Summary", "JD File Path"]
    LINK_IDX   = 4   # 0-based column index
    TITLE_IDX  = 3
    COMPANY_IDX= 2
    SCRAPED_IDX= 10

    existing_rows = []   # [(norm_link, title_lower, company_lower, row_tuple), ...]
    norm_link_seen = {}  # norm_link -> row (保留最新)
    tc_seen = {}          # (title_lower, company_lower) -> norm_link

    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 5 or not row[LINK_IDX]:
                    continue
                norm_link = _normalize_link(str(row[LINK_IDX]))
                title_lc  = str(row[TITLE_IDX] or '').strip().lower()
                comp_lc   = str(row[COMPANY_IDX] or '').strip().lower()
                key_tc    = (title_lc, comp_lc)
                ts        = _parse_scraped_ts(row[SCRAPED_IDX] if len(row) > SCRAPED_IDX else '')

                # 同 norm_link：保留 scraped_at 最新那条
                if norm_link not in norm_link_seen:
                    norm_link_seen[norm_link] = (ts, row)
                else:
                    old_ts, _ = norm_link_seen[norm_link]
                    if ts > old_ts:
                        norm_link_seen[norm_link] = (ts, row)

                # Title+Company 索引（保留 URL 更完整的版本）
                if key_tc not in tc_seen:
                    tc_seen[key_tc] = norm_link
        except Exception as e:
            print(f"[Excel] Load error, creating new: {e}")
            wb = Workbook(); ws = wb.active; ws.title = "AI Jobs"; ws.append(headers)
    else:
        wb = Workbook(); ws = wb.active; ws.title = "AI Jobs"; ws.append(headers)

    print(f"[Excel] {platform}: existing={len(norm_link_seen)} unique jobs")

    # ── 3. 处理新 jobs：去重 + 版本更新 ──────────────────────────
    added   = 0
    updated = 0
    skipped_tc = 0
    rows_to_write = []

    for job in jobs:
        raw_link  = job.get('link', '') or job.get('href', '')
        if not raw_link:
            continue
        norm_link  = _normalize_link(raw_link)
        title_lc   = (job.get('title') or '').strip().lower()
        comp_lc    = (job.get('company') or '').strip().lower()
        key_tc     = (title_lc, comp_lc)
        new_ts     = _parse_scraped_ts(job.get('scraped_at', ''))

        # Title+Company 命中 → 完全跳过（跨平台同一岗位）
        if key_tc in tc_seen:
            skipped_tc += 1
            continue

        # 归一化 URL 已存在 → 比较时间，保留最新
        if norm_link in norm_link_seen:
            old_ts, old_row = norm_link_seen[norm_link]
            if new_ts and old_ts and new_ts > old_ts:
                # 新版本 → 覆盖
                rows_to_write.append((norm_link, title_lc, comp_lc, job, new_ts, True))
                updated += 1
            else:
                # 旧版本或时间相同 → 跳过
                pass
            continue

        # 完全新岗位
        rows_to_write.append((norm_link, title_lc, comp_lc, job, new_ts, False))
        norm_link_seen[norm_link] = (new_ts, None)   # 占位
        tc_seen[key_tc] = norm_link
        added += 1

    # ── 4. 合并：保留 Excel 已有行 + 新增/更新行 ─────────────────
    # 已有行
    merged_rows = [v[1] for v in norm_link_seen.values() if v[1] is not None]
    # 新增/更新行
    for norm_link, title_lc, comp_lc, job, new_ts, is_update in rows_to_write:
        full_jd       = job.get('description', '') or job.get('full_jd', '')
        jd_summary    = extract_jd_summary(full_jd, use_llm=False)  # skip LLM for perf
        jd_file_path  = job.get('jd_file', '')
        source        = job.get('source', platform)

        new_row = [
            0,                          # ID → 稍后填充
            source,
            job.get('company', ''),
            job.get('title', ''),
            raw_link if (job.get('link', '') or job.get('href', '')) == raw_link
               else (job.get('link', '') or job.get('href', '')),
            job.get('location', ''),
            job.get('priority', ''),
            job.get('score', 0),
            job.get('match_reason', ''),
            job.get('post_date', ''),
            job.get('scraped_at', datetime.now().strftime("%Y-%m-%d %H:%M")),
            jd_summary,
            jd_file_path,
        ]
        # 修正：raw_link 变量被循环覆盖，需要用 job 的 link
        new_row[4] = job.get('link', '') or job.get('href', '')
        merged_rows.append(new_row)

    # 排序：Priority(P0>P1>P2>P3>other) + Score desc
    PRI_MAP = {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3}
    def sort_key(r):
        pri = str(r[6] or '')
        return (PRI_MAP.get(pri, 9), -(int(r[7] or 0)))
    merged_rows.sort(key=sort_key)

    # ── 5. 清空 Sheet，重写全部数据 ─────────────────────────────
    # 删除所有数据行（从底往上删避免索引偏移）
    ws.delete_rows(2, ws.max_row)

    for i, row in enumerate(merged_rows, start=1):
        row = list(row)
        row[0] = i       # 重编 ID
        ws.append(row)

    # ── 6. 保存 ─────────────────────────────────────────────────
    try:
        wb.save(excel_path)
        print(f"[Excel] {platform}: +{added} new, ~{updated} updated, "
              f"{skipped_tc} skipped (TC dupe), total={len(merged_rows)} rows")
    except Exception as e:
        print(f"[Excel] Save error: {e}")

# 平台 JD 选择器适配器
JD_SELECTORS = {
    'workday': [
        "[data-automation-id='jobPostingDescription']",
        "[data-automation-id='jobDescription']",
        "[class*='job-posting-description']",
        "[class*='jobDescription']",
        "[class*='description']",
        "[class*='detail']",
        '.job-description'
    ],
    'linkedin': [
        '.jobs-description-content',
        "[class*='job-description']",
        '.jobs-box__html-content'
    ],
    'mokahr': [
        '.job-detail-content',
        "[class*='description']",
        '.job-description'
    ],
    'hays': [
        '.job-description',
        "[data-testid='job-description']"
    ],
    'indeed': [
        '.jobsearch-jobDescriptionText',
        '#jobDescriptionText',
        '.job-description'
    ],
    'taleo': [
        '#requisitionDescriptionInterface',
        '.jobdescription',
        '#job-description'
    ],
    'taleo': [
        '#requisitionDescriptionInterface',
        '[class*="reqDescription"]',
        '[id*="Description"]',
        '.jobdescription',
        '.talentSearchDetailDescription',
        'article.content'
    ],
    'peoplesoft': [
        '[id*="jobdetail"]',
        '[class*="job-detail"]',
        '#jobdetail\\.1',
        '.PSLEVEL1CRDS',
        '[class*="description"]',
        'table.dataGrid td',
        '.pagebody'
    ],
    'pageup': [
        '.job-description',
        '[class*="job-desc"]',
        '[class*="detail-content"]',
        '.content-body',
        '.job-detail-body',
        '#job-description'
    ],
    'jobsdb': [
        '[data-testid="job-description"]',
        '.job-description',
        "[class*='description']",
        '.content-body',
        'article'
    ],
    'default': [
        '.job-description',
        '#job-description',
        '[class*="description"]',
        'article',
        'body'
    ]
}


def get_jd_from_url(jd_page, url, platform='default'):
    """
    统一 JD 抓取函数(Plan C 实现)

    Args:
        jd_page: Playwright page 对象(用于访问 JD 详情页)
        url: JD 详情页 URL
        platform: 平台标识,用于选择正确的选择器(workday/linkedin/mokahr/...)

    Returns:
        str: JD 正文文本(最长 3000 字符)

    特性:
        - 自动适配各平台选择器
        - fallback 机制:试完一个失败自动试下一个
        - 最终兜底:整页 innerText
    """
    if not url or url.startswith('#'):
        return ''

    # 获取平台对应的选择器列表
    selectors = JD_SELECTORS.get(platform, JD_SELECTORS['default'])

    # 尝试每个选择器
    for sel in selectors:
        try:
            if not safe_goto(jd_page, url):
                continue

            random_delay(2, 3)

            # 先尝试关掉 cookie overlay（Workday/中文站点常见）
            for reject_btn in ["button:has-text('拒絕')", "button:has-text('Reject')",
                               "#onetrust-reject-all-handler",
                               "[aria-label*='reject']", "[aria-label*='Reject']"]:
                try:
                    btn = jd_page.query_selector(reject_btn)
                    if btn and btn.is_visible():
                        btn.click()
                        time.sleep(1)
                        break
                except Exception:
                    pass

            # 接受 cookie 兜底（如果拒绝按钮不存在）
            if jd_page.query_selector("[class*='cookie']") or jd_page.query_selector("#onetrust-banner-sdk"):
                for accept_btn in ["button:has-text('接受')", "button:has-text('Accept')",
                                   "#onetrust-accept-btn-handler",
                                   "[aria-label*='accept']"]:
                    try:
                        btn = jd_page.query_selector(accept_btn)
                        if btn and btn.is_visible():
                            btn.click()
                            time.sleep(1)
                            break
                    except Exception:
                        pass

            el = jd_page.query_selector(sel)
            if el:
                text = el.inner_text().strip()
                if len(text) > 100:
                    # 清洗后返回
                    return clean_jd_text(text)
        except Exception:
            continue

    # 最终兜底：整页文本（清洗后返回）
    try:
        if not safe_goto(jd_page, url):
            return ''
        random_delay(2, 3)
        text = jd_page.evaluate("document.body.innerText")
        return clean_jd_text(text) if text else ''
    except Exception:
        return ''
